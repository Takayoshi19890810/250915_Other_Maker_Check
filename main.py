# -*- coding: utf-8 -*-
import os
import re
import io
import time
import json
import unicodedata
from datetime import datetime, timezone, timedelta

import pandas as pd
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

try:
    import google.generativeai as genai
except Exception:
    genai = None

try:
    import jaconv # pip install jaconv
except Exception:
    jaconv = None

# ===== 設定 =====
RELEASE_TAG = "news-latest"
ASSET_NAME = "yahoo_news.xlsx"
SHEET_NAMES = [
    "ホンダ",
    "トヨタ",
    "マツダ",
    "スバル",
    "ダイハツ",
    "スズキ",
    "三菱自動車",
    "日産",
]

def get_keywords() -> list[str]:
    env = os.getenv("NEWS_KEYWORDS")
    if env:
        parts = [p.strip() for p in re.split(r"[,\\n]", env) if p.strip()]
        return parts or SHEET_NAMES
    return SHEET_NAMES

def jst_now():
    return datetime.now(timezone(timedelta(hours=9)))

def jst_str(fmt="%Y/%m/%d %H:%M"):
    return jst_now().strftime(fmt)

def to_hankaku_kana_ascii_digit(s: str) -> str:
    if not s:
        return ""
    s_nfkc = unicodedata.normalize("NFKC", s)
    if jaconv is not None:
        s_nfkc = jaconv.z2h(s_nfkc, kana=True, digit=True, ascii=True)
    return s_nfkc

def normalize_title_for_dup(s: str) -> str:
    if not s:
        return ""
    s = to_hankaku_kana_ascii_digit(s)
    import regex as re_u
    if re_u:
        s = re_u.sub(r'[\p{P}\p{S}\p{Z}\p{Cc}&&[^【】]]+', '', s)
    else:
        dash_chars = r'\\-\\u2212\\u2010\\u2011\\u2012\\u2013\\u2014\\u2015\\uFF0D\\u30FC\\uFF70'
        pattern = (
            r'[\\s"\'\\u201C\\u201D\\u2018\\u2019\\(\\)[\\]{}<>]'
            r'|[、。・,…:;!?！？／/\\\\|＋+＊*.,]'
            r'|[＜＞「」『』《》〔〕［］｛｝（）]'
            r'|[' + dash_chars + r']'
        )
        s = re.sub(pattern, "", s)
    return s

def make_driver() -> webdriver.Chrome:
    opts = Options()
    chrome_path = os.getenv("CHROME_PATH")
    if chrome_path:
        opts.binary_location = chrome_path
    opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--window-size=1280,2000")
    opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)

DATE_RE = re.compile(r"(?:\\d{4}/\\d{1,2}/\\d{1,2}|\\d{1,2}/\\d{1,2})\\s*\\d{1,2}[:：]\\d{2}")

def clean_source_text(text: str) -> str:
    if not text:
        return ""
    t = text
    t = re.sub(r"[（(][^）)]+[）)]", "", t)
    t = DATE_RE.sub("", t)
    t = re.sub(r"^\d+\s*", "", t)
    t = re.sub(r"\s{2,}", " ", t).strip()
    return t

def scrape_yahoo(keyword: str) -> pd.DataFrame:
    driver = make_driver()
    url = (
        f"https://news.yahoo.co.jp/search?p={keyword}"
        f"&ei=utf-8&categories=domestic,world,business,it,science,life,local"
    )
    driver.get(url)
    time.sleep(5)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    driver.quit()

    items = soup.find_all("li", class_=re.compile("sc-1u4589e-0"))
    rows = []
    for li in items:
        try:
            title_tag = li.find("div", class_=re.compile("sc-3ls169-0"))
            link_tag = li.find("a", href=True)
            time_tag = li.find("time")

            title = title_tag.get_text(strip=True) if title_tag else ""
            url = link_tag["href"] if link_tag else ""
            date_str = time_tag.get_text(strip=True) if time_tag else ""

            pub_date = "取得不可"
            if date_str:
                ds = re.sub(r'\\([月火水木金土日]\\)', '', date_str).strip()
                try:
                    dt = datetime.strptime(ds, "%Y/%m/%d %H:%M")
                    pub_date = dt.strftime("%Y/%m/%d %H:%M")
                except ValueError:
                    try:
                        year = jst_now().year
                        dt = datetime.strptime(f"{year}/{ds}", "%Y/%m/%d %H:%M")
                        pub_date = dt.strftime("%Y/%m/%d %H:%M")
                    except ValueError:
                        pub_date = ds

            source = ""
            for sel in [
                "div.sc-n3vj8g-0.yoLqH div.sc-110wjhy-8.bsEjY span",
                "div.sc-n3vj8g-0.yoLqH",
                "span",
                "div",
            ]:
                el = li.select_one(sel)
                if not el:
                    continue
                raw = el.get_text(" ", strip=True)
                txt = clean_source_text(raw)
                if txt and not txt.isdigit():
                    source = txt
                    break

            normalized_title = normalize_title_for_dup(title)

            if title and url:
                rows.append({
                    "タイトル": title, "URL": url, "投稿日": pub_date, "引用元": source or "Yahoo",
                    "取得日時": jst_str(), "検索キーワード": keyword,
                    "ポジネガ": "", "カテゴリ": "", "重複確認用タイトル": normalized_title,
                })
        except Exception:
            continue
    return pd.DataFrame(rows, columns=["タイトル", "URL", "投稿日", "引用元", "取得日時", "検索キーワード", "ポジネガ", "カテゴリ", "重複確認用タイトル"])

def download_existing_book(repo: str, tag: str, asset_name: str, token: str) -> dict[str, pd.DataFrame]:
    empty_cols = ["タイトル", "URL", "投稿日", "引用元", "取得日時", "検索キーワード", "ポジネガ", "カテゴリ", "重複確認用タイトル"]
    dfs: dict[str, pd.DataFrame] = {sn: pd.DataFrame(columns=empty_cols) for sn in SHEET_NAMES}
    if not (repo and tag):
        print("⚠️ download_existing_book: repo/tag が未設定のためスキップ")
        return dfs
    base = "https://api.github.com"
    headers = {"Accept": "application/vnd.github+json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    url_rel = f"{base}/repos/{repo}/releases/tags/{tag}"
    r = requests.get(url_rel, headers=headers)
    print(f"🔎 GET {url_rel} -> {r.status_code}")
    if r.status_code != 200:
        print("⚠️ Releaseが見つからないか、取得に失敗。既存は空として続行します。")
        return dfs
    rel = r.json()
    asset = next((a for a in rel.get("assets", []) if a.get("name") == asset_name), None)
    if not asset:
        print(f"⚠️ Releaseに {asset_name} が存在しません。既存は空として続行します。")
        return dfs
    dl_url = asset.get("browser_download_url")
    if not dl_url:
        print("⚠️ browser_download_url が見つかりません。既存は空として続行します。")
        return dfs
    dr = requests.get(dl_url)
    print(f"⬇️  Download {dl_url} -> {dr.status_code}, {len(dr.content)} bytes")
    if dr.status_code != 200:
        print("⚠️ 既存Excelのダウンロードに失敗。既存は空として続行します。")
        return dfs
    with io.BytesIO(dr.content) as bio:
        try:
            book = pd.read_excel(bio, sheet_name=None, dtype=str)
        except Exception as e:
            print(f"⚠️ 既存Excelの読み込みに失敗: {e}")
            return dfs
    for sn in SHEET_NAMES:
        if sn in book:
            df = book[sn]
            for col in empty_cols:
                if col not in df.columns:
                    df[col] = ""
            dfs[sn] = df[empty_cols].copy()
    return dfs

def save_book_with_format(dfs: dict[str, pd.DataFrame], path: str):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    default_ws = wb.active
    if default_ws:
      wb.remove(default_ws)

    for sheet_name, df in dfs.items():
        ws = wb.create_sheet(title=sheet_name)
        headers = ["タイトル", "URL", "投稿日", "引用元", "取得日時", "検索キーワード", "ポジネガ", "カテゴリ", "重複確認用タイトル"]
        ws.append(headers)
        
        if not df.empty:
            for row in df.itertuples(index=False):
                new_row = list(row)
                try:
                    if pd.notna(row.投稿日):
                        dt_obj = pd.to_datetime(row.投稿日, errors='coerce')
                        if not pd.isna(dt_obj):
                            new_row[2] = dt_obj
                except Exception:
                    pass
                ws.append(new_row)

        max_col = ws.max_column
        max_row = ws.max_row
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        widths = {
            "A": 50, "B": 60, "C": 16, "D": 24, "E": 16,
            "F": 16, "G": 16, "H": 16, "I": 16,
        }
        for col, wdt in widths.items():
            if ws.max_column >= ord(col) - 64:
                ws.column_dimensions[col].width = wdt

        ws.freeze_panes = "A2"

        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = 'yyyy/m/d h:mm'

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)

def classify_with_gemini(dfs: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key or genai is None:
        print("ℹ Gemini分類はスキップ（APIキー未設定 or ライブラリ未インストール）。")
        return dfs

    genai.configure(api_key=api_key)
    model = genai.GenerativeModel("gemini-1.5-flash")
    
    system_prompt = """
あなたは敏腕雑誌記者です。Webニュースのタイトルを以下の規則で厳密に分類してください。

【1】ポジネガ判定（必ず次のいずれか一語のみ）：
- ポジティブ
- ネガティブ
- ニュートラル

【2】記事のカテゴリー判定（最も関連が高い1つだけを選んで出力。並記禁止）：
- 会社：企業の施策や生産、販売台数など。ニッサン、トヨタ、ホンダ、スバル、マツダ、スズキ、ミツビシ、ダイハツの記事の場合は () 付きで企業名を記載。それ以外は「その他」。
- 車：クルマの名称が含まれているもの（会社名だけの場合は車に分類しない）。新型/現行/旧型 + 名称 を () 付きで記載（例：新型リーフ、現行セレナ、旧型スカイライン）。日産以外の車の場合は「車（競合）」と記載。
- 技術（EV）：電気自動車の技術に関わるもの（ただしバッテリー工場建設や企業の施策は含まない）。
- 技術（e-POWER）：e-POWERに関わるもの。
- 技術（e-4ORCE）：4WDや2WD、AWDに関わるもの。
- 技術（AD/ADAS）：自動運転や先進運転システムに関わるもの。
- 技術：上記以外の技術に関わるもの。
- モータースポーツ：F1やラリー、フォーミュラEなど、自動車レースに関わるもの。
- 株式：株式発行や株価の値動き、投資に関わるもの。
- 政治・経済：政治家や選挙、税金、経済に関わるもの。
- スポーツ：野球やサッカー、バレーボールなど自動車以外のスポーツに関わるもの。
- その他：上記に含まれないもの。

【出力要件】
- **JSON配列**のみを返してください（余計な文章や注釈は出力しない）。
- 各要素は次の形式：{"row": 行番号, "sentiment": "ポジティブ|ネガティブ|ニュートラル", "category": "カテゴリ名"}
- 入力の「タイトル」文字列は一切変更しないこと（出力には含めなくて良い）。
""".strip()

    classified_dfs = {}
    for sheet_name, df in dfs.items():
        df_to_classify = df[(df["ポジネガ"] == "") | (df["カテゴリ"] == "")]

        if df_to_classify.empty:
            print(f"ℹ {sheet_name}: 分類対象の行はありません。")
            classified_dfs[sheet_name] = df
            continue

        print(f"✨ {sheet_name}: {len(df_to_classify)}件をGeminiで分類します。")
        df_to_classify = df_to_classify.reset_index(drop=True)

        batch_size = 40
        for start in range(0, len(df_to_classify), batch_size):
            batch = df_to_classify.iloc[start:start + batch_size]
            payload = [{"row": i, "title": t} for i, t in batch.loc[:, ["タイトル"]].itertuples(index=True)]

            try:
                prompt = system_prompt + "\n\n" + json.dumps(payload, ensure_ascii=False, indent=2)
                resp = model.generate_content(prompt)
                text = (resp.text or "").strip()

                import regex as re_u
                m = re_u.search(r'\[.*\]', text, flags=re_u.DOTALL)
                json_text = m.group(0) if m else text
                result = json.loads(json_text)

                for obj in result:
                    try:
                        idx = int(obj.get("row"))
                        sentiment = str(obj.get("sentiment", "")).strip()
                        category = str(obj.get("category", "")).strip()
                        if sentiment and category:
                            df.loc[df_to_classify.index[idx], "ポジネガ"] = sentiment
                            df.loc[df_to_classify.index[idx], "カテゴリ"] = category
                    except Exception as e:
                        print(f"⚠ Gemini応答の解析に失敗: {e}")
            except Exception as e:
                print(f"⚠ Gemini API呼び出しに失敗: {e}")

        classified_dfs[sheet_name] = df
    return classified_dfs

def main():
    keywords = get_keywords()
    print(f"🔎 キーワード一覧: {', '.join(keywords)}")

    token = os.getenv("GITHUB_TOKEN", "")
    repo = os.getenv("GITHUB_REPOSITORY", "")
    dfs_old = download_existing_book(repo, RELEASE_TAG, ASSET_NAME, token)

    dfs_merged: dict[str, pd.DataFrame] = {}
    for kw in keywords:
        df_old = dfs_old.get(kw, pd.DataFrame(columns=["タイトル", "URL", "投稿日", "引用元", "取得日時", "検索キーワード", "ポジネガ", "カテゴリ", "重複確認用タイトル"]))
        df_new = scrape_yahoo(kw)

        df_old['投稿日'] = df_old['投稿日'].astype(str)
        df_new['投稿日'] = df_new['投稿日'].astype(str)
        
        df_all = pd.concat([df_old, df_new], ignore_index=True)
        if not df_all.empty:
            df_all = df_all.dropna(subset=["URL"]).drop_duplicates(subset=["URL"], keep="first")
        dfs_merged[kw] = df_all
        print(f"  - {kw}: 既存 {len(df_old)} 件 + 新規 {len(df_new)} 件 → 合計 {len(df_all)} 件")

    dfs_classified = classify_with_gemini(dfs_merged)

    os.makedirs("output", exist_ok=True)
    out_path = os.path.join("output", ASSET_NAME)
    save_book_with_format(dfs_classified, out_path)

    print(f"✅ Excel出力: {out_path}")
    if repo:
        owner_repo = repo
    else:
        owner_repo = "<OWNER>/<REPO>"
    print(f"🔗 固定DL: https://github.com/{owner_repo}/releases/download/{RELEASE_TAG}/{ASSET_NAME}")

if __name__ == "__main__":
    main()
